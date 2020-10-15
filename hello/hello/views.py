from django.shortcuts import render
from django.http import HttpResponse
from datetime import datetime
import psycopg2
import math
import pandas as pd
from openpyxl import Workbook
import csv
import random

def psql_pdc(query):
    #credenciales PostgreSQL produccion
    connP_P = {
    	'host' : '10.150.1.74',
    	'port' : '5432',
    	'user':'postgres',
    	'password':'cobrando.bi.2020',
    	'database' : 'postgres'}

    #conexion a PostgreSQL produccion
    conexionP_P = psycopg2.connect(**connP_P)
    #print('\nConexión con el servidor PostgreSQL produccion establecida!')
    cursorP_P = conexionP_P.cursor ()
    #ejecucion query telefonos PostgreSQL
    cursorP_P.execute(query)
    anwr = cursorP_P.fetchall()
    
    cursorP_P.close()
    conexionP_P.close()
    return anwr
    
def to_horiz(anwr_P,name,_id):
    #vertical horizontal 
    anwr_P1 = anwr_P.pivot(index=0,columns=1)
    anwr_P1[_id] = anwr_P1.index
    
    col1 = []
    i=0
    for i in range(anwr_P1.shape[1]-1):
        col1.append(name+str(i+1))
    col1.append(_id)
    
    anwr_P1.columns = col1
    
    return anwr_P1

def csv_o(fn,name):
    response = HttpResponse(content_type = "text/csv")
    content = "attachment; filename = %s"%name
    response["Content-Disposition"] = content

    # for j in range(fn.shape[1]):
    #     try:
    #         fn.iloc[:,j] = fn.iloc[:,j].str.decode(encoding='utf-8-sig')
    #         fn.iloc[:,j] = fn.iloc[:,j].str.encode(encoding='utf_16_le')
    #     except:
    #         pass

    fn2 = [tuple(x) for x in fn.values]
    writer = csv.writer(response,delimiter ='|')
    writer.writerow(fn.columns)
    writer.writerows(fn2)

    return response

def excel(fn,name):
    wb = Workbook()
    ws = wb.active

    k = 0
    a = pd.DataFrame(fn.columns)

    for k in range(a.shape[0]):
        ws.cell(row = 1, column = k+1).value = a.iloc[k,0]

    i=0
    j=0

    for i in range(fn.shape[0]):
        for j in range(0,fn.shape[1]):
            try:
                ws.cell(row = i+2, column = j+1).value = fn.iloc[i,j]
            except:
                pass
            
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = %s"%name
    response["Content-Disposition"] = content
    wb.save(response)
    return response

def excel_CV_COL(request):
    today = datetime.now()
    tablename = "CV_Col"+today.strftime("%Y%m%d%H") + ".xlsx"

    with open("./hello/Plantillas/Colp/QueryTel_COL.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Colp/QueryCor_COL.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Colp/QueryDir_COL.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Colp/QueryCV_COL.txt","r") as f4:
        queryP_cons = f4.read()
        
    with open("./hello/Plantillas/Colp/QueryCiu_COL.txt","r") as f6:
        queryP_Ciu = f6.read()

    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infCi = to_horiz(anwr_Ci,'town',"deudor_id")


    df = df.rename(columns={0:'rownumber',
                            1:'obligacion_id',
                            2:'deudor_id',
                            3:'unico',
                            4:'estado',
                            5:'tipo_cliente',
                            6:'nombre',
                            7:'producto',
                            8:'initial_bucket',
                            9:'ciudad',
                            10:'sucursal',
                            11:'tipo_prod',
                            12:'dias_mora_inicial',
                            13:'dias_mora_actual',
                            14:'rango_mora_inicial',
                            15:'rango_mora_final',
                            16:'rango',
                            17:'suma_pareto',
                            18:'rango_pareto',
                            19:'fcast',
                            20:'fdesem',
                            21:'vrdesem',
                            22:'saldo_total_inicial',
                            23:'saldo_total_actual',
                            24:'saldo_capital_inicial',
                            25:'saldo_capital_actual',
                            26:'saldo_vencido_inicial',
                            27:'saldo_vencido_actual',
                            28:'pagomin',
                            29:'fultpago',
                            30:'vrultpago',
                            31:'agencia',
                            32:'tasainter',
                            33:'feultref',
                            34:'ultcond',
                            35:'fasigna',
                            36:'eqasicampana',
                            37:'diferencia_pago',
                            38:'pago_preliminar',
                            39:'pago_cliente',
                            40:'min',
                            41:'tarifa',
                            42:'honorarios',
                            43:'perfil_mes_4',
                            44:'perfil_mes_3',
                            45:'perfil_mes_2',
                            46:'perfil_mes_1',
                            47:'fecha_primer_gestion',
                            48:'fecha_ultima_gestion',
                            49:'perfil_mes_actual',
                            50:'contactabilidad',
                            51:'ultimo_alo',
                            52:'descod1',
                            53:'descod2',
                            54:'asesor',
                            55:'fecha_gestion',
                            56:'telefono_mejor_gestion',
                            57:'mejorgestionhoy',
                            58:'asesor_indicador_hoy',
                            59:'repeticion',
                            60:'llamadas',
                            61:'sms',
                            62:'correos',
                            63:'gescall',
                            64:'visitas',
                            65:'whatsapp',
                            66:'no_contacto',
                            67:'total_gestiones',
                            68:'telefono_positivo',
                            69:'marcaciones_telefono_positivo',
                            70:'ultima_marcacion_telefono_positivo',
                            71:'fec_creacion_ult_compromiso',
                            72:'fec_pactada_ult_compromiso',
                            73:'valor_acordado_ult_compromiso',
                            74:'asesor_ult_compromiso',
                            75:'cantidad_acuerdos_mes',
                            76:'estado_acuerdo',})

    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return excel(fn,tablename)


def csv_CV_Claro(request):
    today = datetime.now()
    tablename = "CV_Claro" + today.strftime("%Y%m%d%H") + ".csv"

    with open("./hello/Plantillas/Claro/QueryTel_Claro.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Claro/QueryCor_Claro.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Claro/QueryCV_Claro.txt","r") as f4:
        queryP_cons = f4.read()
        
    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    yanwr = psql_pdc(queryP_cons)

    #dataframes
    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    df = pd.DataFrame(yanwr)

    anwr_P1 = to_horiz(anwr_P,'phone','deudor_id')

    #renombrar campos correos
    anwr_C = anwr_C.rename(columns={
                            
                            0:'deudor_id',
                            1:'mail0',
                            2:'mail1'})

    anwr_C1 = anwr_C.drop_duplicates(subset=['deudor_id'])

    #renombrar campos CV
    df = df.rename(columns={0:'rownumber',
    1:'deudor_id',
    2:'obligacion_id',
    3:'nombredelcliente',
    4:'estado',
    5:'tipo_cliente',
    6:'unico',
    7:'crmorigen',
    8:'potencialmark',
    9:'prepotencialmark',
    10:'writeoffmark',
    11:'dias_mora',
    12:'segmento_bpo',
    13:'rango_bpo',
    14:'tipo',
    15:'fecha_de_vencimiento',
    16:'min_cliente',
    17:'valorscoring',
    18:'numeroreferenciadepago',
    19:'monto_inicial',
    20:'monto_ini_cuenta',
    21:'porcentaje_descuento',
    22:'valor_descuento',
    23:'valor_a_pagar',
    24:'deuda_real',
    25:'valor_pago',
    26:'saldo_pendiente',
    27:'fecha_pago',
    28:'fecha_compromiso',
    29:'fecha_pago_compromiso',
    30:'valor_compromiso',
    31:'estado_acuerdo',
    32:'ind_m4',
    33:'ind_m3',
    34:'ind_m2',
    35:'ind_m1',
    36:'fecha_primer_gestion',
    37:'fecha_ultima_gestion',
    38:'indicador',
    39:'phone',
    40:'asesor',
    41:'fecha_gestion',
    42:'contactabilidad',
    43:'indicador_hoy',
    44:'repeticion',
    45:'llamadas',
    46:'sms',
    47:'correos',
    48:'gescall',
    49:'whatsapp',
    50:'visitas',
    51:'no_contacto',
    52:'total_gestiones',
    53:'telefono_positivo',
    54:'fec_ultima_marcacion'})

    #a = fn[fn.obligacion_id == '9876510000211227']
    #i=0
    #lin = ['no_contacto_mes_actual','gescall_mes_actual','tel_mes_actual','tel_positivo']
    #for i in lin:
    #    df[i].fillna(0,inplace=True)
    #    df[i] = df[i].apply(lambda x: round(x))
    #    df[i] = df[i].astype('str')

    fn = pd.merge(df,anwr_P1,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,anwr_C1,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return csv_o(fn,tablename)

def csv_CV_CarP(request):
    today = datetime.now()
    tablename = "CV_CarP" + today.strftime("%Y%m%d%H")  + ".csv"

    with open("./hello/Plantillas/CarP/QueryTel_CarP.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/CarP/QueryCor_CarP.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/CarP/QueryDir_CarP.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/CarP/QueryCV_CarP.txt","r") as f4:
        queryP_cons = f4.read()

    with open("./hello/Plantillas/CarP/QueryCiu_CarP.txt","r") as f5:
        queryP_Ciu = f5.read()


    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infCi = to_horiz(anwr_Ci,'town',"deudor_id")

    #renombrar campos CV
    df = df.rename(columns={0:'deudor_id',
                            1:'unico',
                            2:'nombre',
                            3:'obligacion',
                            4:'obligacion_17',
                            5:'tipo_cliente',
                            6:'sucursal_final',
                            7:'zona',
                            8:'ano_castigo',
                            9:'saldo_k_pareto_mes_vigente',
                            10:'intereses',
                            11:'honorarios_20',
                            12:'saldo_total_mes_vigente',
                            13:'saldo_total_pareto_mes_vigente_',
                            14:'saldokpareto',
                            15:'rango_k_pareto',
                            16:'interesespareto',
                            17:'honorariospareto',
                            18:'porcentaje_k_del_total',
                            19:'porcentaje_intereses_del_total',
                            20:'porcentaje_honorarios_del_total',
                            21:'rango_k_porcentaje',
                            22:'capital_20_porciento',
                            23:'dias_mora_acumulado',
                            24:'marca_juridica_cliente',
                            25:'focos',
                            26:'valor_pago',
                            27:'ultima_fecha_pago',
                            28:'estado_cliente_mes_anterior',
                            29:'valor_compromiso',
                            30:'fecha_compromiso',
                            31:'fecha_pactada_compromiso',
                            32:'asesor_compromiso',
                            33:'ind_m4',
                            34:'ind_m3',
                            35:'ind_m2',
                            36:'ind_m1',
                            37:'fecha_primer_gestion',
                            38:'fecha_ultima_gestion',
                            39:'indicador',
                            40:'telefono_mejor_gestion',
                            41:'asesor_mejor_gestion',
                            42:'fecha_gestion',
                            43:'contactabilidad',
                            44:'indicador_hoy',
                            45:'repeticion',
                            46:'llamadas',
                            47:'sms',
                            48:'correos',
                            49:'gescall',
                            50:'whatsapp',
                            51:'visitas',
                            52:'no_contacto',
                            53:'total_gestiones',
                            54:'telefono_positivo',
                            55:'fec_ultima_marcacion',
                            56:'investigacion_de_bienes'})


    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return csv_o(fn,tablename)                

def csv_CV_FalaJ(request):
    today = datetime.now()
    tablename = "CV_FalJ"+today.strftime("%Y%m%d%H") + ".csv"

    with open("./hello/Plantillas/Fala/QueryTel_Fal.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Fala/QueryCor_Fal.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Fala/QueryDir_Fal.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Fala/QueryCV_FalJ.txt","r") as f4:
        queryP_cons = f4.read()

    with open("./hello/Plantillas/Fala/QueryPa_Fal.txt","r") as f5:
        queryP_Pa = f5.read()

    with open("./hello/Plantillas/Fala/QueryRe_Fal.txt","r") as f6:
        queryP_PR = f6.read()
        
    with open("./hello/Plantillas/Fala/QueryCiu_Fal.txt","r") as f7:
        queryP_Ciu = f7.read()

    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrPa = psql_pdc(queryP_Pa)
    anwrR = psql_pdc(queryP_PR)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Pa = pd.DataFrame(anwrPa)
    anwr_R = pd.DataFrame(anwrR)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infR = to_horiz(anwr_R,'referencia',"deudor_id")
    infCi = to_horiz(anwr_Ci,'town',"deudor_id")

    try:
        infP = to_horiz(anwr_Pa,'pago','obligacion_id')
        
        if infP.shape[1] > 4:
            tipos = infP.dtypes.to_frame()
            tipos['index'] = range(len(tipos))
            tipos = tipos.set_index('index')
        
        su = []
        for n in range(len(tipos)):
            if str(tipos.iloc[n,0]) == 'float64':
                su.append(n)
            else:
                pass
        
        infP1 = infP[['pago1','pago2']]
        infP1['pago3'] = infP.iloc[:,2:max(su)+1].sum(axis = 1)
        infP1['obligacion_id'] = infP.index
        infP = infP1
        
        i=0
        lin = ['pago1','pago2','pago3']
        for i in lin:
            infP[i].fillna(0,inplace=True)
            infP[i] = infP[i].apply(lambda x: round(x))
            infP[i] = '$' + infP[i].astype('str')
    except:
        pass

    #renombrar campos CV
    df = df.rename(columns={0:'idcbpo',
    1:'tipo_producto_asignacion',
    2:'grupo',
    3:'cartera',
    4:'tipo_cliente',
    5:'unico',
    6:'unico_pro',
    7:'obligacion_id',
    8:'deudor_id',
    9:'nombre',
    10:'producto',
    11:'saldototal',
    12:'saldo_pareto',
    13:'segmentacion',
    14:'peso',
    15:'alturamora_hoy',
    16:'rango',
    17:'dias_mora',
    18:'vencto',
    19:'indicador_mejor_gestion',
    20:'total_gestiones',
    21:'fecha_ultima_gestion',
    22:'asesor_mejor_gestion',
    23:'fecha_compromiso',
    24:'fecha_pago_compromiso',
    25:'valor_compromiso',
    26:'asesor',
    27:'estado_acuerdo',
    28:'dias_mora_pagos',
    29:'valor_pago',
    30:'fecha_pago',
    31:'pendiente',
    32:'pago_total',
    33:'nvo_status',
    34:'status_refresque',
    35:'nvo_status_refresque',
    36:'dias_mora_refresque',
    37:'pendiente_mas_gastos',
    38:'vencida_mas_gastos',
    39:'gastos_mora',
    40:'gastos_cv',
    41:'porcentaje_gasto',
    42:'valor_a_mantener_sin_gxc',
    43:'cv8',
    44:'cv9',
    45:'cv10',
    46:'cv11',
    47:'cv12',
    48:'restructuracion',
    49:'valor_restruc',
    50:'pagominimo_actual',
    51:'pagominimo_anterior',
    52:'periodo_actual',
    53:'periodo_anterior',
    54:'cuota36',
    55:'cuota48',
    56:'cuota60',
    57:'cuota72',
    58:'proyectada_cargue',
    59:'aplica_ajuste',
    60:'fecha',
    61:'diferencia',
    62:'porcentaje_saldo_total',
    63:'x',
    64:'valor',
    65:'porcentaje_participacion',
    66:'ind_m4',
    67:'ind_m3',
    68:'ind_m2',
    69:'ind_m1',
    70:'fecha_primer_gestion',
    71:'telefono_mejor_gestion',
    72:'fecha_gestion',
    73:'contactabilidad',
    74:'indicador_hoy',
    75:'repeticion',
    76:'llamadas',
    77:'sms',
    78:'correos',
    79:'gescall',
    80:'whatsapp',
    81:'visitas',
    82:'no_contacto',
    83:'telefono_positivo',
    84:'fec_ultima_marcacion',
    85:'lista_robinson'})


    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infR,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    if 'infP' in locals():
    #    Cruce pagos
        fn = pd.merge(fn,infP,on = ["obligacion_id"]\
                    ,how = "left",indicator = False)
    #   ordenamiento
        lt = fn.columns.tolist()
        lt = lt[:29] + lt[(infP.shape[1]-1)*-1:] + lt[29:fn.shape[1]-(infP.shape[1]-1)]
        fn = fn[lt]

    return csv_o(fn,tablename)

def csv_CV_FalaC(request):
    today = datetime.now()
    tablename = "CV_FalC"+today.strftime("%Y%m%d%H") + ".csv"

    with open("./hello/Plantillas/Fala/QueryTel_Fal.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Fala/QueryCor_Fal.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Fala/QueryDir_Fal.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Fala/QueryCV_FalC.txt","r") as f4:
        queryP_cons = f4.read()

    with open("./hello/Plantillas/Fala/QueryPa_Fal.txt","r") as f5:
        queryP_Pa = f5.read()

    with open("./hello/Plantillas/Fala/QueryRe_Fal.txt","r") as f6:
        queryP_PR = f6.read()
        
    with open("./hello/Plantillas/Fala/QueryCiu_Fal.txt","r") as f7:
        queryP_Ciu = f7.read()

    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrPa = psql_pdc(queryP_Pa)
    anwrR = psql_pdc(queryP_PR)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Pa = pd.DataFrame(anwrPa)
    anwr_R = pd.DataFrame(anwrR)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infR = to_horiz(anwr_R,'referencia',"deudor_id")
    infCi = to_horiz(anwr_Ci,'town',"deudor_id")
    try:
        infP = to_horiz(anwr_Pa,'pago','obligacion_id')
        
        if infP.shape[1] > 4:
            tipos = infP.dtypes.to_frame()
            tipos['index'] = range(len(tipos))
            tipos = tipos.set_index('index')
        
        su = []
        for n in range(len(tipos)):
            if str(tipos.iloc[n,0]) == 'float64':
                su.append(n)
            else:
                pass
        
        infP1 = infP[['pago1','pago2']]
        infP1['pago3'] = infP.iloc[:,2:max(su)+1].sum(axis = 1)
        infP1['obligacion_id'] = infP.index
        infP = infP1
        
        i=0
        lin = ['pago1','pago2','pago3']
        for i in lin:
            infP[i].fillna(0,inplace=True)
            infP[i] = infP[i].apply(lambda x: round(x))
            infP[i] = '$' + infP[i].astype('str')
    except:
        pass

    #renombrar campos CV
    df = df.rename(columns={0:'idcbpo',
    1:'tipo_producto_asignacion',
    2:'grupo',
    3:'cartera',
    4:'tipo_cliente',
    5:'unico',
    6:'unico_pro',
    7:'obligacion_id',
    8:'deudor_id',
    9:'nombre',
    10:'producto',
    11:'saldototal',
    12:'saldo_pareto',
    13:'segmentacion',
    14:'peso',
    15:'alturamora_hoy',
    16:'rango',
    17:'dias_mora',
    18:'vencto',
    19:'indicador_mejor_gestion',
    20:'total_gestiones',
    21:'fecha_ultima_gestion',
    22:'asesor_mejor_gestion',
    23:'fecha_compromiso',
    24:'fecha_pago_compromiso',
    25:'valor_compromiso',
    26:'asesor',
    27:'estado_acuerdo',
    28:'dias_mora_pagos',
    29:'valor_pago',
    30:'fecha_pago',
    31:'pendiente',
    32:'pago_total',
    33:'nvo_status',
    34:'status_refresque',
    35:'nvo_status_refresque',
    36:'dias_mora_refresque',
    37:'pendiente_mas_gastos',
    38:'vencida_mas_gastos',
    39:'gastos_mora',
    40:'gastos_cv',
    41:'porcentaje_gasto',
    42:'valor_a_mantener_sin_gxc',
    43:'cv1',
    44:'cv2',
    45:'cv3',
    46:'cv4',
    47:'cv5',
    48:'cv6',
    49:'cv7',
    50:'cv8',
    51:'cv9',
    52:'cv10',
    53:'cv11',
    54:'cv12',
    55:'restructuracion',
    56:'valor_restruc',
    57:'pagominimo_actual',
    58:'pagominimo_anterior',
    59:'periodo_actual',
    60:'periodo_anterior',
    61:'cuota36',
    62:'cuota48',
    63:'cuota60',
    64:'cuota72',
    65:'proyectada_cargue',
    66:'aplica_ajuste',
    67:'fecha',
    68:'diferencia',
    69:'porcentaje_saldo_total',
    70:'x',
    71:'valor',
    72:'porcentaje_participacion',
    73:'ind_m4',
    74:'ind_m3',
    75:'ind_m2',
    76:'ind_m1',
    77:'fecha_primer_gestion',
    78:'telefono_mejor_gestion',
    79:'fecha_gestion',
    80:'contactabilidad',
    81:'indicador_hoy',
    82:'repeticion',
    83:'llamadas',
    84:'sms',
    85:'correos',
    86:'gescall',
    87:'whatsapp',
    88:'visitas',
    89:'no_contacto',
    90:'telefono_positivo',
    91:'fec_ultima_marcacion',
    92:'lista_robinson'})


    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infR,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    if 'infP' in locals():
    #    Cruce pagos
        fn = pd.merge(fn,infP,on = ["obligacion_id"]\
                    ,how = "left",indicator = False)
    #   ordenamiento
        lt = fn.columns.tolist()
        lt = lt[:27] + lt[(infP.shape[1]-1)*-1:] + lt[27:fn.shape[1]-(infP.shape[1]-1)]
        fn = fn[lt]

    return csv_o(fn,tablename)

def csv_CV_Sant(request):
    today = datetime.now()
    tablename = "CV_San"+today.strftime("%Y%m%d%H") + ".csv"

    with open("./hello/Plantillas/Sant/QueryTel_San.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Sant/QueryCor_San.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Sant/QueryDir_San.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Sant/QueryCV_San.txt","r") as f4:
        queryP_cons = f4.read()
        
    with open("./hello/Plantillas/Sant/QueryCiu_San.txt","r") as f5:
        queryP_Ciu = f5.read()
        
    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infCi = to_horiz(anwr_Ci,'town',"deudor_id")

    #renombrar campos CV
    df = df.rename(columns={0:'entidad',
                            1:'abogado',
                            2:'fecha_desembolso',
                            3:'fecha_corte',
                            4:'solicitud',
                            5:'obligacion_id',
                            6:'deudor_id',
                            7:'unico',
                            8:'nombre',
                            9:'capital',
                            10:'dias_mora',
                            11:'rango_mora',
                            12:'saldo_capital_pareto',
                            13:'rango_pareto',
                            14:'tomo_encuesta',
                            15:'aplica_alivio',
                            16:'fecha_proximo_pago_alivio',
                            17:'debitos',
                            18:'rango_cierre_m4',
                            19:'rango_cierre_m3',
                            20:'rango_cierre_m2',
                            21:'rango_cierre_m1',
                            22:'repeticion',
                            23:'llamadas',
                            24:'sms',
                            25:'correos',
                            26:'gescall',
                            27:'whatsapp',
                            28:'visitas',
                            29:'no_contacto',
                            30:'total_gestiones',
                            31:'fecha_primer_gestion',
                            32:'fecha_ultima_gestion',
                            33:'ultimo_alo',
                            34:'ind_m4',
                            35:'ind_m3',
                            36:'ind_m2',
                            37:'ind_m1',
                            38:'ind_mes_actual',
                            39:'fec_ind_mes_actual',
                            40:'tel_ind_mes_actual',
                            41:'asesor_ind_mes_actual',
                            42:'contactabilidad',
                            43:'asesor_ind_hoy',
                            44:'ind_hoy',
                            45:'ultima_marcacion_tel_pos',
                            46:'telefono_positivo',
                            47:'fecha_compromiso',
                            48:'fecha_pago_compromiso',
                            49:'valor_compromiso',
                            50:'calificacion',
                            51:'fechas_probables_pago',
                            52:'id_protocolo',
                            53:'canal_protocolo',
                            54:'texto_protocolo'})
            
    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    lt = fn.columns.tolist()
    lt = lt[:52] + lt[-(inf.shape[1] + infC.shape[1] + infD.shape[1] -1):] + lt[52:55]
    fn = fn[lt]
    
    return csv_o(fn,tablename)

def csv_CV_Pop(request):
    today = datetime.now()
    tablename = "CV_Pop" + today.strftime("%Y%m%d%H") + '.csv'

    with open("./hello/Plantillas/Pop/QueryTel_Pop.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Pop/QueryCor_Pop.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Pop/QueryDir_Pop.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Pop/QueryCV_Pop.txt","r") as f4:
        queryP_cons = f4.read()

    with open("./hello/Plantillas/Pop/QueryPa_Pop.txt","r") as f5:
        queryP_Pa = f5.read()
        
    with open("./hello/Plantillas/Pop/QueryCiu_Pop.txt","r") as f6:
        queryP_Ciu = f6.read()
        
    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrPa = psql_pdc(queryP_Pa)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Pa = pd.DataFrame(anwrPa)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infCi = to_horiz(anwr_Ci,'town',"deudor_id")
    try:
        infP = to_horiz(anwr_Pa,'pago','llave')
    except:
        pass

    #renombrar campos CV
    df = df.rename(columns={0:'tipo_bpo',
    1:'llave',
    2:'entidad',
    3:'unico',
    4:'nombre_prod',
    5:'lin_descripcion',
    6:'deudor_id',
    7:'nombre_cliente',
    8:'tipo_cliente',
    9:'estado_actual_cartera',
    10:'obligacion_id',
    11:'inicio_corte',
    12:'fecha_desembolso',
    13:'ciclo',
    14:'dias_en_mora_inicial',
    15:'dias_en_mora_final',
    16:'valor_compromiso',
    17:'fecha_creacion_compromiso',
    18:'fecha_pago_compromiso',
    19:'asesor_compromiso',
    20:'pagos_acumulados',
    21:'cantidad_pagos',
    22:'rango_mora_inicial',
    23:'rango_mora_final',
    24:'estado',
    25:'valmora',
    26:'valmora_pareto',
    27:'capital_inicial',
    28:'saltotalpareto',
    29:'pago_minimo',
    30:'rango_saltotal',
    31:'asignacion_inicial',
    32:'wasis_banco',
    33:'fecha_de_retiro',
    34:'tipo_cliente',
    35:'ind_m4',
    36:'ind_m3',
    37:'ind_m2',
    38:'ind_m1',
    39:'ind_mejor_gestion',
    40:'fec_mejor_gestion',
    41:'tel_mejor_gestion',
    42:'asesor_mejor_gestion',
    43:'contactability',
    44:'ind_mejor_gestion_hoy',
    45:'asesor_mejor_gestion_hoy',
    46:'fecha_primer_gestion',
    47:'fecha_ultima_gestion',
    48:'repeticion',
    49:'llamadas',
    50:'sms',
    51:'correos',
    52:'gescall',
    53:'whatsapp',
    54:'visitas',
    55:'no_contacto',
    56:'total_gestiones',
    57:'primer_alo',
    58:'ultimo_alo',
    59:'fec_ult_marc_tel_pos',
    60:'tel_positivo',
    61:'casa_inicial',
    62:'casa_actual',
    63:'fecha_retiro_casa'})
            
    # inf["deudor_id"] = "1" + inf["deudor_id"]
    # infC["deudor_id"] = "1" + infC["deudor_id"]
    # infD["deudor_id"] = "1" + infD["deudor_id"]
    # infCi["deudor_id"] = "1" + infCi["deudor_id"]

    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    if 'infP' in locals():
    #    Cruce pagos
        fn = pd.merge(fn,infP,on = ["llave"]\
                    ,how = "left",indicator = False)

    #   ordenamiento
        lt = fn.columns.tolist()
        lt = lt[:20] + lt[(infP.shape[1]-1)*-1:] + lt[20:fn.shape[1]-(infP.shape[1]-1)]
        fn = fn[lt]

    return csv_o(fn,tablename)

def csv_CV_Dav(request):
    today = datetime.now()
    tablename = "CV_Dav"+today.strftime("%Y%m%d%H") + ".csv"

    with open("./hello/Plantillas/Davi/QueryTel_Dav.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Davi/QueryCor_Dav.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Davi/QueryDir_Dav.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Davi/QueryCV_Dav.txt","r") as f4:
        queryP_cons = f4.read()

    with open("./hello/Plantillas/Davi/QueryCiu_Dav.txt","r") as f5:
        queryP_Ciu = f5.read()
        
    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infCi = to_horiz(anwr_Ci,'town',"deudor_id")

    #renombrar campos CV
    df = df.rename(columns={0:'tipo_de_cliente',
    1:'cartera',
    2:'deudor_id',
    3:'obligacion_id',
    4:'nueva_obligacion',
    5:'nombre_cliente',
    6:'macroportafolio',
    7:'producto',
    8:'detalle_producto',
    9:'unico',
    10:'exposicion_cliente',
    11:'franja_mora',
    12:'dias',
    13:'saldo_capital',
    14:'pago_minimo',
    15:'saldo_mora',
    16:'saldo_pareto',
    17:'rango_pareto',
    18:'fecha_desembolso',
    19:'plazo',
    20:'plazo_restante',
    21:'no_reestructuraciones',
    22:'tasa_de_interes',
    23:'ciclo_pago',
    24:'cuota_prod',
    25:'dias_mora_hoy',
    26:'rxm_proyectado',
    27:'pago_minimo_us',
    28:'cabeza_para_mora_30',
    29:'cabeza_para_mora_60',
    30:'resultado_productoproy',
    31:'resultado_clienteproy',
    32:'rxm_proy_max',
    33:'status_resultado_accion',
    34:'rango_llamadas_efectivas',
    35:'saldo_total_actual',
    36:'cia_actual_min',
    37:'cia_actual',
    38:'segmentacion_sfc_inicial',
    39:'obligacion_def',
    40:'fecha_cambio_no_obligacion',
    41:'valor_dcto',
    42:'valor_minimo_pagar',
    43:'estrategia',
    44:'dia',
    45:'congelado',
    46:'descod1',
    47:'descod2',
    48:'macro_banco',
    49:'tipo_garantia',
    50:'precastigos',
    51:'tombolizacion',
    52:'saldo_contable',
    53:'contable_pareto',
    54:'salto_factura',
    55:'fecha_creacion_compromiso',
    56:'fecha_pago_compromiso',
    57:'asesor_compromiso',
    58:'valor_compromiso',
    59:'valor_pago',
    60:'fecha_pago',
    61:'macrofactura',
    62:'estado_compromiso',
    63:'ind_m4',
    64:'ind_m3',
    65:'ind_m2',
    66:'ind_m1',
    67:'ind_mejor_gestion',
    68:'fec_mejor_gestion',
    69:'tel_mejor_gestion',
    70:'asesor_mejor_gestion',
    71:'contactability',
    72:'ind_mejor_gestion_hoy',
    73:'asesor_mejor_gestion_hoy',
    74:'alo',
    75:'fecha_primer_gestion',
    76:'fecha_ultima_gestion',
    77:'repeticion',
    78:'llamadas',
    79:'sms',
    80:'correos',
    81:'gescall',
    82:'whatsapp',
    83:'visitas',
    84:'no_contacto',
    85:'total_gestiones',
    86:'ultimo_alo',
    87:'fec_ult_marc_tel_pos',
    88:'tel_positivo'})
            
    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return csv_o(fn,tablename)

def csv_CV_Cod(request):
    today = datetime.now()
    tablename = "CV_Cod" + today.strftime("%Y%m%d%H") + '.csv'

    with open("./hello/Plantillas/Cod/QueryTel_Cod.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Cod/QueryCor_Cod.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Cod/QueryDir_Cod.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Cod/QueryCV_Cod.txt","r") as f4:
        queryP_cons = f4.read()

    with open("./hello/Plantillas/Cod/QueryCiu_Cod.txt","r") as f5:
        queryP_Ciu = f5.read()
        
    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrCi = psql_pdc(queryP_Ciu)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Ci = pd.DataFrame(anwrCi)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone','identificacion')
    infC = to_horiz(anwr_C,'mail','identificacion')
    infD = to_horiz(anwr_D,'address','identificacion')
    infCi = to_horiz(anwr_Ci,'town','identificacion')

    #renombrar campos CV
    df = df.rename(columns={0:'identificacion',
    1:'id_cliente',
    2:'nombre',
    3:'producto',
    4:'cliente',
    5:'dias_inicial',
    6:'franja_mora_inicial',
    7:'dias_actual',
    8:'franja_mora_actual',
    9:'mora',
    10:'inteses',
    11:'inteses_mo',
    12:'saldo_mora',
    13:'total',
    14:'saldo_capital',
    15:'saldo_pareto',
    16:'rango_saldo',
    17:'ci',
    18:'tipo_cliente',
    19:'nuevo_antiguo',
    20:'alivios',
    21:'pago_cliente',
    22:'fechapago',
    23:'paga_paga_o_no',
    24:'pago_real',
    25:'actualizacion_diaria',
    26:'valor_compromiso',
    27:'fecha_compromiso',
    28:'fecha_pago_compromiso',
    29:'asesor_compromiso',
    30:'estado_acuerdo',
    31:'ind_m4',
    32:'ind_m3',
    33:'ind_m2',
    34:'ind_m1',
    35:'indicador_mejor_gestion',
    36:'fecha_mejor_gestion',
    37:'telefono_mejor_gestion',
    38:'asesor_mejor_gestion',
    39:'tipo_contacto',
    40:'fecha_primer_gestion',
    41:'fecha_ultima_gestion',
    42:'indicador_mejor_gestion_hoy',
    43:'asesor_mejor_gestion_hoy',
    44:'repeticion',
    45:'llamadas',
    46:'sms',
    47:'correos',
    48:'gescall',
    49:'whatsapp',
    50:'visitas',
    51:'no_contacto',
    52:'total_gestiones',
    53:'fecha_ultimo_alo',
    54:'telefono_positivo',
    55:'fecha_telefono_positivo'})
            
    fn = pd.merge(df,inf,on = ['identificacion']\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ['identificacion']\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ['identificacion']\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infCi,on = ['identificacion']\
                ,how = "left",indicator = False)
    
    return csv_o(fn,tablename)

def csv_GesD_Claro(request):
    today = datetime.now()
    tablename = "Ges_Claro"+today.strftime("%Y%m%d%H")+'.csv'

    with open("./hello/Plantillas/Claro/QueryDia_Claro.txt","r") as f1:
        queryP_PT = f1.read()
        
    anwr = psql_pdc(queryP_PT)
    fn = pd.DataFrame(anwr)

    #renombrar campos CV
    fn = fn.rename(columns={0:'deudor_id',
                            1:'dia',
                            2:'indicador',
                            3:'repeticion',
                            4:'llamadas',
                            5:'sms',
                            6:'correos',
                            7:'gescall',
                            8:'whatsapp',
                            9:'no_contacto',
                            10:'fecha_gestion',
                            11:'visitas',
                            12:'phone',
                            13:'asesor',
                            14:'descod01',
                            15:'descod02'})
    
    return csv_o(fn,tablename)

def csv_GesD_Davi(request):
    today = datetime.now()
    tablename = "Ges_Davi"+today.strftime("%Y%m%d%H")+'.csv'

    with open("./hello/Plantillas/Davi/QueryDia_Dav.txt","r") as f1:
        queryP_PT = f1.read()
        
    anwr = psql_pdc(queryP_PT)
    fn = pd.DataFrame(anwr)

    #renombrar campos CV
    fn = fn.rename(columns={0:'deudor_id',
                            1:'dia',
                            2:'indicador',
                            3:'repeticion',
                            4:'llamadas',
                            5:'sms',
                            6:'correos',
                            7:'gescall',
                            8:'whatsapp',
                            9:'no_contacto',
                            10:'fecha_gestion',
                            11:'visitas',
                            12:'phone',
                            13:'asesor',
                            14:'descod01',
                            15:'descod02'})
    
    return csv_o(fn,tablename)

def Rep_Chat(request,num):
    lista = ['cbpo_bogota'
    ,'cbpo_carteraok'
    ,'cbpo_claro'
    ,'cbpo_davivienda'
    ,'cbpo_falabella'
    ,'cbpo_popular'
    ,'cbpo_progresa'
    ,'cbpo_propia'
    ,'cbpo_qnt']
    #credenciales PostgreSQL produccion
    connP_P = {
    	'host' : '10.150.1.74',
    	'port' : '5432',
    	'user':'postgres',
    	'password':'cobrando.bi.2020',
    	'database' : 'postgres'}
    
    queryP_data = """
    select distinct agent_id
    from """+lista[num]+""".wolk_chats
    where chat_date between date_trunc('month',current_date) and current_date
    order by agent_id desc;
    """
    queryP_data1 = """
    select distinct chat_id
    from """+lista[num]+""".wolk_chats
    where chat_date between date_trunc('month',current_date) and current_date
    and agent_id = %s
    order by chat_id desc;
    """
    queryP_data2 = """
    select chat_id ,from_msg ,from_ ,to_ ,"time" ,msg 
    from """+lista[num]+""".wolk_conv
    where chat_id = %s
    order by time;
    """
    conexionP_P = psycopg2.connect(**connP_P)
    cursorP_P = conexionP_P.cursor ()
    cursorP_P.execute(queryP_data)
    anwr = cursorP_P.fetchall()
    
    #i = 2
    for i in range(len(anwr)):
        agent = str(anwr[i][0])
        cursorP_P.execute(queryP_data1 % agent)
        anwr1 = cursorP_P.fetchall()
        if len(anwr1)-1 > 0:
            chat = str(anwr1[random.randrange(0,len(anwr1)-1)][0])
        else:
            chat = str(anwr1[0][0])
        cursorP_P.execute(queryP_data2 % chat)
        if 'anwr2' not in locals():
            anwr2 = pd.DataFrame(cursorP_P.fetchall())
        else:
            anwr2 = anwr2.append(pd.DataFrame(cursorP_P.fetchall()))
    anwr2 = anwr2.rename(columns={0:'chat_id',
                            1:'from_msg',
                            2:'emisor',
                            3:'receptor',
                            4:'fecha',
                            5:'msg'})
    return csv_o(anwr2,lista[num]+'.csv')