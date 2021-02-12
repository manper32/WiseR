from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from rest_framework import generics
from file_app.serializers import FileSerializer
from file_app.serializers import GestionSerializer
from file_app.serializers import TareasSerializer
from file_app.serializers import VicidialPauseSerializer
from file_app.serializers import CampaignListSerializer
from file_app.serializers import AuxIndicativosSerializer
from file_app.serializers import HabeasDataSerializer
from file_app.models import Tipificaciones, Codigos, NombreRama
from django.db.models import Avg, Max, Min, Sum, Count, Sum
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from pyexcel_xlsx import get_data
from file_app.models import Gestiones
from file_app.models import Tareas
from file_app.models import Asignaciones
from file_app.models import IndicadoresGeneral
from file_app.models import VicidialPause
from file_app.models import Unidad
from file_app.models import TipificacionesHerramientas
from file_app.models import UsuariosWiser
from file_app.models import CampaingList
from file_app.models import Habeasdata
from file_app.models import Habeasdata
from file_app.models import AuxIndicativos
from file_app.models import Telefonos
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from django.http import JsonResponse
import jxmlease
import psycopg2
import requests
import pandas as pd
import mysql.connector
import json
import random

def excel(fn,name):
    wb = Workbook()
    ws = wb.active

    k = 0
    a = pd.DataFrame(fn.columns)

    for k in range(a.shape[0]):
        ws.cell(row = 1, column = k+1).value = a.iloc[k,0]

    i=0
    j=0

    for i in range(fn.shape[0]):
        for j in range(0,fn.shape[1]):
            try:
                ws.cell(row = i+2, column = j+1).value = fn.iloc[i,j]
            except:
                pass
            
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = %s"%name
    response["Content-Disposition"] = content
    wb.save(response)
    return response

def SMS_Mas(tel='',msg='',lms='false',fls='false',pmu='true'):

        # credentials Vicidial
        connM = {
            'user' : 'Api_UGYHS',
            'psw' : 'MP2JBEKOUY',
            'url' : 'https://sms.masivapp.com/SmsHandlers/sendhandler.ashx'}
        
        args = {'action': 'sendmessage',
                'username' : connM.get('user'),
                'password' : connM.get('psw'),
                'recipient' : '57'+tel,
                'messagedata' : msg,
                'longMessage' : lms,
                'flash' : fls,
                'premium' : pmu}
        
        b = requests.get(connM.get('url').replace('\n','')
                        ,params=args).text
        
        b = [jxmlease.parse(b)['response']['action'].get_cdata(),
        jxmlease.parse(b)['response']['data']['acceptreport']['statuscode'].get_cdata(),
        jxmlease.parse(b)['response']['data']['acceptreport']['statusmessage'].get_cdata(),
        jxmlease.parse(b)['response']['data']['acceptreport']['messageid'].get_cdata(),
        jxmlease.parse(b)['response']['data']['acceptreport']['recipient'].get_cdata(),
        jxmlease.parse(b)['response']['data']['acceptreport']['messagetype'].get_cdata(),
        jxmlease.parse(b)['response']['data']['acceptreport']['messagedata'].get_cdata()]

        return b

def AL_Vici(list_id,campaign,active,descr,list_name,local_call_time,numip):

    if int(numip) > 200:
        # n = '150'
        connV = {
        'server' : '10.150.1.'+numip,
        'agc' : 'vicidial',
        'user' : '6666',
        'psw' : 'cobrando2012',
        'url' : 'http://{0}/{1}/non_agent_api.php'}
    else:
        # n = '152'
        connV = {
            'server' : '10.150.1.'+numip,
            'agc' : 'vicidial',
            'user' : 'secetina',
            'psw' : '1233692646',
            'url' : 'http://{0}/{1}/non_agent_api.php'
            }
        
    args = {
        'source':'test',
        'function':'add_list',
        'user':connV.get('user'),
        'pass':connV.get('psw'),
        'list_id':list_id,
        'list_name':list_name,
        'campaign_id':campaign,
        'active':active,
        'list_description':descr,
        # 'custom_fields_copy':'202115'
    }


    # url
    # with open("/home/manuel/Documentos/WiseR/Vicidial/Templates/AL_URL.txt","r") as f1:
        # al_url = f1.read()
        
    b = pd.DataFrame([requests.get(connV.get('url').\
                     format(connV.get('server'),
                            connV.get('agc')),params=args).text.split('|')])
    return b

def ALe_Vici(phone,list_id,Vendor_lead,numip):

    if int(numip) == 206:
        # n = '150'
        connV = {
        'server' : '10.150.1.'+numip,
        'agc' : 'vicidial',
        'user' : 'soporte',
        'psw' : 'Bogota1234',
        'url' : 'http://{0}/{1}/non_agent_api.php?source=test&user={2}&pass={3}&function=add_lead&phone_number=9{4}&list_id={5}&vendor_lead_code={6}'}
    elif int(numip) == 209:
        # n = '150'
        connV = {
        'server' : '10.150.1.'+numip,
        'agc' : 'vicidial',
        'user' : 'soporte',
        'psw' : 'Bogota1234',
        'url' : 'http://{0}/{1}/non_agent_api.php?source=test&user={2}&pass={3}&function=add_lead&phone_number={4}&list_id={5}&vendor_lead_code={6}'}
    else:
        # n = '152'
        connV = {
        'server' : '10.150.1.'+numip,
        'agc' : 'vicidial',
        'user' : 'secetina',
        'psw' : '1233692646',
        'url' : 'http://{0}/{1}/non_agent_api.php?source=test&user={2}&pass={3}&function=add_lead&phone_number={4}&list_id={5}&vendor_lead_code={6}'}

    # url
    # with open("/home/manuel/Documentos/WiseR/Vicidial/Templates/ALe_URL.txt","r") as f1:
        # ale_url = f1.read()
        
    b = pd.DataFrame([requests.get(connV.get('url').format(connV.get('server'),
                                    connV.get('agc'),
                                    connV.get('user'),
                                    connV.get('psw'),
                                    phone,
                                    list_id,
                                    Vendor_lead).replace('\n','')).text.split('|')])
    return b

# Tipificaciones
@method_decorator(csrf_exempt,name='dispatch')
class FileTipi(APIView):

    parser_classes = (MultiPartParser, FormParser)
    def post(self, request, *args, **kwargs):
        file_serializer = FileSerializer(data=request.data)
        if file_serializer.is_valid():
            # file_serializer.save()
            
            file_obj = request.data['file']
            data1 = get_data(file_obj)
            col = list(data1.keys())
            data = pd.DataFrame(data1[col[0]][1:],columns=data1[col[0]][0])
            try:
                data2 = pd.DataFrame(data1[col[1]][1:],columns=data1[col[1]][0])
            except:
                pass
            # print(data1[col[1]][2])

            cols = ['codigo1',
                    'codigo2',
                    'codigo3',
                    'codigo4',
                    'codigo5',
                    'codigo6',
                    'codigo7',
                    'codigo8',
                    'codigo9',
                    'codigo10',
                    'prioridad',
                    'indicador']

            #credenciales PostgreSQL produccion
            connP_P = {
                'host' : '10.150.1.77',
                'port' : '5432',
                'user':'bi',
                'password':'juanitoMeToco2020',
                'database' : 'login'}

            query = """select id,indicador
                    from public.indicadores_general
                    order by indicador ;"""
            
            # tipificaciones

            tipi = pd.DataFrame(columns=cols)
            tipi_n = pd.DataFrame(columns=cols)
            unq = pd.DataFrame()
            names = []
            n = 0


            for i in data:
                if data[i].dtype == 'object':
                    
                    names.append(i.lower())
                    tipi[i.lower()] = data[i].str.upper().str.strip()
                    unq = pd.concat([unq,
                                    data[i].str.upper().str.strip().drop_duplicates()\
                                        .reset_index(drop=True)],
                                    ignore_index=True,
                                    axis=1)
                    unq.columns = names
                    if i.lower().find('codigo') > -1:
                        
                        names.append(i.lower()+'_n')
                        unq = pd.concat([unq,pd.DataFrame(range(len(data[i].str.upper().str.strip()\
                                    .drop_duplicates().reset_index(drop=True))))],
                                    ignore_index=True,
                                    axis=1)
                        unq.columns = names
                        if i.lower()[-1] != '0':
                            unq[i.lower()+'_n'] = unq[i.lower()+'_n'] + (int(i.lower()[-1])*100)
                        else:
                            unq[i.lower()+'_n'] = unq[i.lower()+'_n'] + (int(i.lower()[-2:])*100)
                elif data[i].dtype == 'int64':
                    tipi[i.lower()] = data[i]
                    
            tipi.fillna(0,inplace=True)


            for i in unq:
                if i.lower().find('codigo') > -1 and i.lower().find('_n') == -1 and n == 0:
                    tipi_n[i] = pd.merge(tipi[[i]],
                                    unq[[i,i+'_n']],
                                    on = [i],
                                    how = "left",
                                    indicator = False)[i+'_n'].astype(int)
                    n += 1
                elif i.lower().find('codigo') > -1 and i.lower().find('_n') == -1:
                    tipi_n[[i]] = pd.merge(tipi[[i]],
                                                unq[[i,i+'_n']].fillna(0),# merge correccion con el try
                                                on = [i],
                                                how = "left",
                                                indicator = False)[i+'_n'].astype(int)
            
            tipi_n['indicador'] = pd.merge(tipi['indicador'],
                                pd.read_sql(query,psycopg2.connect(**connP_P)),
                                on = ['indicador'],
                                how='left',
                                indicator = False)['id']
                        
            for i in tipi:
                if tipi[i].dtype == 'int64':
                    tipi_n[i] = tipi[i]

            for i in range(len(tipi_n)):
                try:    
                    Tipificaciones.objects.using(request.data.get('remark')).create(codigo01=tipi_n.loc[i,['codigo1']],
                                                                    codigo02=tipi_n.loc[i,['codigo2']],
                                                                    codigo03=tipi_n.loc[i,['codigo3']],
                                                                    codigo04=tipi_n.loc[i,['codigo4']],
                                                                    codigo05=tipi_n.loc[i,['codigo5']],
                                                                    codigo06=tipi_n.loc[i,['codigo6']],
                                                                    codigo07=tipi_n.loc[i,['codigo7']],
                                                                    codigo08=tipi_n.loc[i,['codigo8']],
                                                                    codigo09=tipi_n.loc[i,['codigo9']],
                                                                    codigo10=tipi_n.loc[i,['codigo10']],
                                                                    prioridad=tipi_n.loc[i,['prioridad']],
                                                                    indicador=tipi_n.loc[i,['indicador']],
                                                                    unidad=self.kwargs['unidad'])
                except:
                    pass

            for i in cols:
                if i.lower().find('codigo') > -1 and i in unq.columns:
                    for j in range(len(unq[[i,i+'_n']].dropna())):
                        Codigos.objects.using(request.data.get('remark')).create(descripcion=unq.loc[j,[i]][0],
                                                                                codigo=unq.loc[j,[i+'_n']][0],
                                                                                unidad=self.kwargs['unidad'])
            
            if 'data2' in locals():
                for i in range(len(data2)):
                    NombreRama.objects.using(request.data.get('remark')).create(id=i+1,
                                                                                nombre=data2.loc[i,['nombre cliente']][0],
                                                                                unidad=self.kwargs['unidad'])

            return Response(file_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# envio SMS
class FileSMS(APIView):

    parser_classes = (MultiPartParser, FormParser)
    def post(self, request, unidad):
        file_serializer = FileSerializer(data=request.data)
        if file_serializer.is_valid():
            # file_serializer.save()

            file_obj = request.data['file']
            data = get_data(file_obj)
            col = list(data.keys())
            data = pd.DataFrame(data[col[0]][1:],columns=data[col[0]][0])
            
            try:
                asignacion = Asignaciones.objects.using(request.data.get('remark')).get(estado = True,
                                                                                    unidad = unidad)
            except:
                return Response({'error':'No coincide la unidad'}, status=status.HTTP_400_BAD_REQUEST)

            tipificacion = TipificacionesHerramientas.objects.using(request.data.get('remark')).get(
                herramienta='SMS'
            )
            # asignacion = Asignaciones.objects.using(request.data.get('remark')).get(estado = True,
                                                                        # unidad = unidad)

            tarea = Tareas.objects.using(request.data.get('remark')).create(
                tarea_id = Tareas.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('tarea_id')).get('tarea_id__max')+1,
                unidad_id = unidad,
                registros = len(data),
                clientes = len(data.cedula.drop_duplicates()),
                obligaciones = 0,
                tipo = 'SMS')
            # print(tarea)
            
            for i in range(len(data)):
                
                Gestiones.objects.using(request.data.get('remark')).create(
                gestion_id = Gestiones.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('gestion_id')).get('gestion_id__max')+1
                ,tarea_id = tarea.tarea_id
                ,usuario_id = 'SMS_BACK'
                ,deudor_id = data['cedula'][i]
                ,asignacion_id = asignacion.asignacion_id
                ,telefono = data['telefono'][i]
                ,canal = 'SMS'
                ,id_tipificacion = tipificacion.id
                ,descripcion = data['mensaje'][i])

                SMS_Mas(str(data['telefono'][i]),
                            data['mensaje'][i])            

            return Response(file_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# retorno de llamadas
class RetornoLlamadas(APIView):
    def post(self, request, *args, **kwargs):
        file_serializer = FileSerializer(data=request.data)
        if file_serializer.is_valid():

            file_obj = request.data['file']
            data = get_data(file_obj)
            col = list(data.keys())
            data = pd.DataFrame(data[col[0]][1:],columns=data[col[0]][0])

            if self.kwargs['unidad'] == 0:
                prefijo = '9'
            elif self.kwargs['unidad'] != 0:
                prefijo = Unidad.objects.using('public').get(id=self.kwargs['unidad']).prefijo
                if prefijo == None:
                    return Response({'error':'Añadir el prefijo a la unidad'}, status=status.HTTP_400_BAD_REQUEST)
            # print(prefijo)

            url = 'http://10.150.1.206/vicidial/non_agent_api.php'
            
            Plist = []
            for i in range(len(data)):
                args = {
                'source' : 'test',
                'user' : 'soporte',
                'pass' : 'Bogota1234',
                'function' : 'add_lead',
                'phone_number' : str(prefijo)+str(data.loc[i,['telefono']][0]),
                'list_id' : '20200811001',
                'vendor_lead_code' : str(data.loc[i,['cedula']][0])
                }
                # print(args)
                Plist.append(requests.get(url,params=args).status_code)

            tarea = Tareas.objects.using(request.data.get('remark')).create(
                tarea_id = Tareas.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('tarea_id')).get('tarea_id__max')+1,
                unidad_id = self.kwargs['unidad'],
                registros = len(data),
                clientes = len(data.cedula.drop_duplicates()),
                obligaciones = 0,
                tipo = 'RETURN',
                list_id = '20200811001')
            
            return Response(file_serializer.data, status=status.HTTP_200_OK)

# Envio de EMAIL
class FileEmail(APIView):
    def post(self, request, *args, **kwargs):
        file_serializer = FileSerializer(data=request.data)
        if file_serializer.is_valid():

            file_obj = request.data['file']
            data = get_data(file_obj)
            col = list(data.keys())
            data = pd.DataFrame(data[col[0]][1:],columns=data[col[0]][0])

            try:
                asignacion = Asignaciones.objects.using(request.data.get('remark')).get(estado = True,
                                                                                    unidad = self.kwargs['unidad'])
            except:
                return Response({'error':'No coincide la unidad'}, status=status.HTTP_400_BAD_REQUEST)

            tipificacion = TipificacionesHerramientas.objects.using(request.data.get('remark')).get(
                herramienta='EMAIL'
            )

            tarea = Tareas.objects.using(request.data.get('remark')).create(
                tarea_id = Tareas.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('tarea_id')).get('tarea_id__max')+1,
                unidad_id = self.kwargs['unidad'],
                registros = len(data),
                clientes = len(data.cedula.drop_duplicates()),
                obligaciones = 0,
                tipo = 'EMAIL')

            for i in range(len(data)):
                Gestiones.objects.using(request.data.get('remark')).create(
                gestion_id = Gestiones.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('gestion_id')).get('gestion_id__max')+1
                ,tarea_id = tarea.tarea_id
                ,usuario_id = 'EMAIL_BACK'
                ,deudor_id = data['cedula'][i]
                ,asignacion_id = asignacion.asignacion_id
                ,canal = 'EMAIL'
                ,id_tipificacion = tipificacion.id
                ,descripcion = data['mensaje'][i]
                ,nom_contacto_tercero = data['correo'][i])
            # print(data)
            

            return Response(file_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Envio de GesCall
class FileGesCall(APIView):
    def post(self, request, *args, **kwargs):
        file_serializer = FileSerializer(data=request.data)
        if file_serializer.is_valid():

            file_obj = request.data['file']
            data = get_data(file_obj)
            col = list(data.keys())
            data = pd.DataFrame(data[col[0]][1:],columns=data[col[0]][0])
            
            try:
                asignacion = Asignaciones.objects.using(request.data.get('remark')).get(estado = True,
                                                                                    unidad = self.kwargs['unidad'])
            except:
                return Response({'error':'No coincide la unidad'}, status=status.HTTP_400_BAD_REQUEST)

            tipificacion = TipificacionesHerramientas.objects.using(request.data.get('remark')).get(
                herramienta='GESCALL'
            )

            tarea = Tareas.objects.using(request.data.get('remark')).create(
                tarea_id = Tareas.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('tarea_id')).get('tarea_id__max')+1,
                unidad_id = self.kwargs['unidad'],
                registros = len(data),
                clientes = len(data.cedula.drop_duplicates()),
                obligaciones = 0,
                tipo = 'GESCALL')            

            for i in range(len(data)):
                Gestiones.objects.using(request.data.get('remark')).create(
                gestion_id = Gestiones.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('gestion_id')).get('gestion_id__max')+1
                ,tarea_id = tarea.tarea_id
                ,usuario_id = 'GESCALL'
                ,deudor_id = data['cedula'][i]
                ,asignacion_id = asignacion.asignacion_id
                ,telefono = data['telefono'][i]
                ,canal = 'GESCALL'
                ,id_tipificacion = tipificacion.id
                ,descripcion = data['texto'][i])
            # print(data)

            return Response(file_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# creacion tareas call
class FileCreacionTarea(APIView):

    parser_classes = (MultiPartParser, FormParser)
    def post(self, request, numip, unidad, campaign, name, callf):
        file_serializer = FileSerializer(data=request.data)
        if file_serializer.is_valid():
            # file_serializer.save()

            file_obj = request.data['file']
            data = get_data(file_obj)
            col = list(data.keys())
            data = pd.DataFrame(data[col[0]][1:],columns=data[col[0]][0])

            now = datetime.now()
            # print(now.strftime("%Y%m%d%H%M%S"))
            a = AL_Vici(now.strftime("%Y%m%d%H%M%S"),
                        campaign.replace(r'\n',''),
                        'Y',#active,
                        '--BLANK--',#descr,
                        name,#list_name,
                        '24hours',#local_call_time,
                        str(numip))
            # print(Gestiones.objects.using(request.data.get('remark'))\
                    # .all().aggregate(Max('tarea_id')).get('tarea_id__max')+2)

            if not a.iloc[0,0].find('SUCCESS: add_list LIST HAS BEEN ADDED') > -1:
                return Response({'result' : a.iloc[0,0]},status = status.HTTP_400_BAD_REQUEST)

            for i in range(len(data)):
                a = ALe_Vici(data['telefono'][i],
                        now.strftime("%Y%m%d%H%M%S"),
                        data['cedula'][i],
                        str(numip))
                if not a.iloc[0,0].find('SUCCESS: add_lead LEAD HAS BEEN ADDED') > -1:
                    return Response({'result' : a.iloc[0,0]},status = status.HTTP_400_BAD_REQUEST)
                
            if callf == 1:
                tipo = 'CALLF'
            elif callf == 0:
                tipo = 'CALL'
            else:
                return Response({'error' : 'callf es binario'},status = status.HTTP_400_BAD_REQUEST)

            tarea = Tareas.objects.using(request.data.get('remark')).create(
                tarea_id = Tareas.objects.using(request.data.get('remark'))\
                    .all().aggregate(Max('tarea_id')).get('tarea_id__max')+1,
                unidad_id = unidad,
                registros = len(data),
                clientes = len(data.cedula.drop_duplicates()),
                obligaciones = 0,
                tipo = tipo,
                nombre = name,
                list_id = now.strftime("%Y%m%d%H%M%S"))

            return Response(file_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Habeas data
class CrearHabeasData(APIView):
    def post(self, request, *args, **kwargs):
        habeas_data=request.data.get('habeas_data')
        deudor_id=request.data.get('deudor_id')
        telefono=request.data.get('telefono')

        if habeas_data == None:
            return Response({'error':'habeas_data requerido'},status=status.HTTP_400_BAD_REQUEST)
        elif deudor_id == None:
            return Response({'error':'deudor_id requerido'},status=status.HTTP_400_BAD_REQUEST)
        elif telefono == None:
            return Response({'error':'telefono requerido'},status=status.HTTP_400_BAD_REQUEST)

        Habeasdata.objects.using(self.kwargs['db']).create(
            deudor_id=deudor_id,
            habeas_data=habeas_data,
            telefono=telefono)

        if habeas_data == True and len(str(telefono)) == 7:
            # departamento = request.data.get('departaento')
            ciudad = request.data.get('ciudad')
            Telefonos.objects.using(self.kwargs['db']).create(
                telefono = telefono,
                deudor_id = deudor_id,
                telefono_tipo = 'Fijo',
                telefono_estado = True,
                ciudad = ciudad,
                indicativo = AuxIndicativos.objects.using('public').get(ciudad=ciudad).indicativo
            )
        elif habeas_data == True and len(str(telefono)) == 10:
            Telefonos.objects.using(self.kwargs['db']).create(
                telefono = telefono,
                deudor_id = deudor_id,
                telefono_tipo = 'Movil',
                telefono_estado = True,
            )
        else:
            return Response({'error':'numero tiene que ser de 7 o 10 digitos'},status=status.HTTP_400_BAD_REQUEST)
        return Response({'result':'Telefono agregado'},status=status.HTTP_201_CREATED)

# Habeas data
class ConsultaHabeasData(APIView):
    def get(self, request, *args, **kwargs):
        queryset = Habeasdata.objects.using(self.kwargs['db'])\
            .filter(fecha_registro__range= [self.kwargs['li'],self.kwargs['ls']]).values()
        b = pd.DataFrame(list(queryset))
        if b.empty:
            return Response({'result':'Intervalo sin registros'},status=status.HTTP_200_OK)
        b['habeas_data'] = b.habeas_data.replace({
            0:'Rechazado',
            1:'Aceptado'
            })
        return excel(b,'HabeasData'+self.kwargs['db']+'.xlsx')

# consulta gestiones historicas
class ConsultaGestion(APIView):
    def get(self, request, *args, **kwargs):
        queryset1 = pd.DataFrame(list(Gestiones.objects.using(self.kwargs['db'])\
            .filter(deudor_id=self.kwargs['deudor_id']).values(
                'gestion_id',
                'tarea_id',
                'gestion_fecha',
                'usuario_id',
                'deudor_id',
                'asignacion_id',
                'telefono',
                'canal',
                'id_tipificacion',
                'descripcion',
                'nom_contacto_tercero',
                'tel_adicional',
                'ciudad_tel_adicional',)))
        queryset3 = pd.DataFrame(list(Tipificaciones.objects.using(self.kwargs['db']).all()\
            .values('id','indicador')))
        queryset2 = pd.DataFrame(list(IndicadoresGeneral.objects.using('public').all()\
                .values('id','indicador')))
        queryset4 = pd.DataFrame(list(UsuariosWiser.objects.using('public').all()\
                .values('id','nombre')))

        if queryset1.empty:
            return Response({'result':'Sin gestiones anteriores'}, status=status.HTTP_200_OK)

        merge1 = pd.merge(queryset1
                            ,queryset3
                            ,how = "left"
                            ,left_on='id_tipificacion'
                            ,right_on='id'
                            ,indicator=False).drop(['id'],axis=1)
        merge2 = pd.merge(merge1
                            ,queryset2
                            ,how = "left"
                            ,left_on='indicador'
                            ,right_on='id'
                            ,indicator=False).drop(['id'],axis=1)
        merge3 = pd.merge(merge2
                            ,queryset4
                            ,how = "left"
                            ,left_on=merge2.usuario_id
                            ,right_on=queryset4.id.astype(str)
                            ,indicator=False).drop(['id'],axis=1).to_json(orient='records')

        return JsonResponse(json.loads(merge3),safe=False)

# consulta gestiones historicas
class ConsultaTareaCall(generics.ListCreateAPIView):
    def get(self, request, *args, **kwargs):
        queryset = pd.DataFrame(list(Tareas.objects.using(self.kwargs['db'])\
            .filter(tarea_fecha_creacion__gte= datetime.now().replace(day=1
                                                                            ,hour=0
                                                                            ,minute=0
                                                                            ,second=0
                                                                            ,microsecond=0)
                    ,tipo='CALL').values()))

        if queryset.empty:
            return Response({'result':'Sin tareas Previas'}, status=status.HTTP_200_OK)

                #credenciales MySQL120
        connM = {
            'host' : '10.150.1.'+self.kwargs['numip'],
            'user':'desarrollo',
            'password':'soportE*8994',
            'database' : 'asterisk'
        }
        query = """SELECT
t2.list_name
,t1.list_id
,sum(case when t1.status in ('MS','SMS','NE''NC','BZ','MST','CO','PSC','FLL','FL','ME','FAS','TRA','PU','NC') then 1 else 0 end ) no_contacto
,sum(case when t1.status in ('PS','PSC') then 1 else 0 end ) presion
,sum(case when t1.status in ('SG','YP') then 1 else 0 end ) seguimiento
,sum(case when t1.status in ('CP','CR') then 1 else 0 end ) compromiso
,sum(case when t1.status in ('NG') then 1 else 0 end ) negociacion
,sum(case when t1.status in ('MS','SMS','MST') then 1 else 0 end ) mensaje
FROM asterisk.vicidial_list t1
left join asterisk.vicidial_lists t2
on t1.list_id = t2.list_id 
-- where t1.list_id = 29012021123456;
-- WHERE t2.list_name = 'ADELANTOS NVO 2901'
where CAST(t1.modify_date as date) >= SUBDATE(CURRENT_DATE(), DAYOFMONTH(CURRENT_DATE()) - 1)
GROUP by t1.list_id
order by t1.list_id;
        """
        b= pd.merge(queryset
                    ,pd.read_sql(query,mysql.connector.Connect(**connM))
                    ,how = "left"
                    ,on = 'list_id'
                    ,indicator=False).to_json(orient='records')
        # print(b)
        return JsonResponse(json.loads(b),safe=False)
        # return Response({'status':'ok'}, status=status.HTTP_201_CREATED)

# consulta tarea SMS
class ConsultaTareaSMS(generics.ListCreateAPIView):
    def get_queryset(self):
        queryset = Tareas.objects.using(self.kwargs['db'])\
            .filter(tarea_fecha_creacion__gte= datetime.now().replace(day=1
                                                                            ,hour=0
                                                                            ,minute=0
                                                                            ,second=0
                                                                            ,microsecond=0)
                                                                            ,tipo='SMS')
        return queryset
    serializer_class = TareasSerializer

# consulta tarea EMAIL
class ConsultaTareaEMAIL(generics.ListCreateAPIView):
    def get_queryset(self):
        queryset = Tareas.objects.using(self.kwargs['db'])\
            .filter(tarea_fecha_creacion__gte= datetime.now().replace(day=1
                                                                            ,hour=0
                                                                            ,minute=0
                                                                            ,second=0
                                                                            ,microsecond=0)
                                                                            ,tipo='EMAIL')
        return queryset
    serializer_class = TareasSerializer

# consulta tarea Gescall
class ConsultaTareaGesCall(generics.ListCreateAPIView):
    def get_queryset(self):
        queryset = Tareas.objects.using(self.kwargs['db'])\
            .filter(tarea_fecha_creacion__gte= datetime.now().replace(day=1
                                                                            ,hour=0
                                                                            ,minute=0
                                                                            ,second=0
                                                                            ,microsecond=0)
                                                                            ,tipo='GESCALL')
        return queryset
    serializer_class = TareasSerializer

# consulta vicidial campaigns
class ConsultaCampaignList(generics.ListCreateAPIView):
    def get_queryset(self):
        queryset = CampaingList.objects.using(self.kwargs['db']).all()
        return queryset
    serializer_class = CampaignListSerializer

# consulta vicidial pause
class ConsultaVicidialPause(generics.ListCreateAPIView):
    def get_queryset(self):
        queryset = VicidialPause.objects.using('public').all()
        return queryset
    serializer_class = VicidialPauseSerializer

# consulta tareas Return
class ConsultaTareasReturn(generics.ListCreateAPIView):
    def get_queryset(self):
        queryset = Tareas.objects.using(self.kwargs['db']).filter(
            tipo='RETURN',
            unidad_id=self.kwargs['unidad'],
            tarea_fecha_creacion__gte=datetime.now().replace(day=1
                                                            ,hour=0
                                                            ,minute=0
                                                            ,second=0
                                                            ,microsecond=0))
        return queryset
    serializer_class = TareasSerializer

# consulta tareas Return
class ExcelTareasReturn(APIView):
    def get(self, request, *args, **kwargs):
        queryset = Tareas.objects.using(self.kwargs['db']).filter(
            tipo='RETURN',
            unidad_id=self.kwargs['unidad'],
            tarea_fecha_creacion__gte=datetime.now().replace(day=1
                                                            ,hour=0
                                                            ,minute=0
                                                            ,second=0
                                                            ,microsecond=0)).values()
        b = pd.DataFrame(list(queryset))
        
        return excel(b,'Return'+self.kwargs['db']+'.xlsx')

# consulta suma de herramienta
class ConsultaTareasSum(APIView):
    def get(self, request, *args, **kwargs):
        queryset = Tareas.objects.using(self.kwargs['db'])\
                .filter(tipo=self.kwargs['tipo'],
                        tarea_fecha_creacion__gte=datetime.now().replace(day=1
                                                                    ,hour=0
                                                                    ,minute=0
                                                                    ,second=0
                                                                    ,microsecond=0))\
                .values('tipo')\
                .annotate(rcount=Sum('registros'))
                
        # print(queryset)

        return Response(queryset, status=status.HTTP_201_CREATED)

# consulta vicidial campaigns
class ConsultaAuxIndicativos(generics.ListCreateAPIView):
    def get_queryset(self):
        queryset = AuxIndicativos.objects.using('public').all()
        return queryset
    serializer_class = AuxIndicativosSerializer

class WolkvoxRepChat(APIView):
    def get(self, request, *args, **kwargs):

        lista = [
            'cbpo_bogota'
            ,'cbpo_carteraok'
            ,'cbpo_claro'
            ,'cbpo_davivienda'
            ,'cbpo_falabella'
            ,'cbpo_popular'
            ,'cbpo_progresa'
            ,'cbpo_propia'
            ,'cbpo_qnt']
        #credenciales PostgreSQL produccion
        connP_P = {
            'host' : '10.150.1.74',
            'port' : '5432',
            'user':'postgres',
            'password':'cobrando.bi.2020',
            'database' : 'postgres'
            }
        
        queryP_data = """
        select distinct agent_id
        from """+lista[self.kwargs['num']]+""".wolk_chats
        where chat_date between date_trunc('month',current_date) and current_date
        order by agent_id desc;
        """
        queryP_data1 = """
        select distinct chat_id
        from """+lista[self.kwargs['num']]+""".wolk_chats
        where chat_date between date_trunc('month',current_date) and current_date
        and agent_id = {0}
        order by chat_id desc;
        """
        queryP_data2 = """
        select chat_id ,from_msg ,from_ ,to_ ,"time" ,msg 
        from """+lista[self.kwargs['num']]+""".wolk_conv
        where chat_id = {0}
        order by time;
        """
        conexionP_P = psycopg2.connect(**connP_P)
        cursorP_P = conexionP_P.cursor ()
        cursorP_P.execute(queryP_data)
        anwr = cursorP_P.fetchall()
        
        #i = 2
        for i in range(len(anwr)):
            agent = str(anwr[i][0])
            cursorP_P.execute(queryP_data1.format(agent))
            anwr1 = cursorP_P.fetchall()
            if len(anwr1)-1 > 0:
                chat = str(anwr1[random.randrange(0,len(anwr1)-1)][0])
            else:
                chat = str(anwr1[0][0])
            cursorP_P.execute(queryP_data2.format(chat))
            if 'anwr2' not in locals():
                anwr2 = pd.DataFrame(cursorP_P.fetchall())
            else:
                anwr2 = anwr2.append(pd.DataFrame(cursorP_P.fetchall()))
        anwr2 = anwr2.rename(columns={0:'chat_id',
                                1:'from_msg',
                                2:'emisor',
                                3:'receptor',
                                4:'fecha',
                                5:'msg'}).to_json(orient='records')

        return JsonResponse(json.loads(anwr2),safe=False)
        # return Response({'ok':'ok'}, status=status.HTTP_201_CREATED)